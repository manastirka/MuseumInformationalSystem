#!/usr/bin/env python3
"""
Import Bird Ringing (Prstenovanje ptica) Data from Excel to SQLite Database
This script processes the comprehensive bird ringing dataset for the Natural History Museum
"""

import sqlite3
import openpyxl
from datetime import datetime
import os

# Database file path
DB_PATH = 'data/bird_ringing.db'

def create_database():
    """Create the bird ringing database and table."""
    # Ensure data directory exists
    os.makedirs('data', exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Create table with all relevant fields from the Excel file
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bird_ringing (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ring_number INTEGER,
            color_ring TEXT,
            species TEXT,
            age INTEGER,
            sex TEXT,
            location TEXT,
            coordinates TEXT,
            coordinate_accuracy TEXT,
            date TEXT,
            time TEXT,
            metal_ring_position INTEGER,
            plastic_ring TEXT,
            catching_method TEXT,
            bait TEXT,
            manipulation TEXT,
            status TEXT,
            clutch_size TEXT,
            pullus_age TEXT,
            wing_length TEXT,
            third_primary_feather TEXT,
            mass TEXT,
            molting TEXT,
            back_claw TEXT,
            bill_length TEXT,
            bill_measurement_method TEXT,
            head_length TEXT,
            tarsus TEXT,
            tarsus_measurement_method TEXT,
            tail_length TEXT,
            fat_deposits TEXT,
            pectoral_muscle TEXT,
            incubation_patch TEXT,
            alula TEXT,
            carpal_feathers TEXT,
            sex_determination TEXT,
            protected_areas TEXT,
            ringer TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Create indexes for better search performance
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_species ON bird_ringing(species)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_location ON bird_ringing(location)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_date ON bird_ringing(date)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_ringer ON bird_ringing(ringer)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_ring_number ON bird_ringing(ring_number)")

    conn.commit()
    conn.close()
    print(f"✓ Database created: {DB_PATH}")

def clean_value(value):
    """Clean and convert values from Excel."""
    if value is None:
        return None

    # Convert datetime objects to string
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')

    # Convert to string and strip whitespace
    value_str = str(value).strip()

    # Return None for empty strings
    if value_str == '' or value_str == 'None':
        return None

    return value_str

def import_data():
    """Import data from Excel file to SQLite database."""
    print("📂 Loading Excel file...")
    wb = openpyxl.load_workbook('Prstenovanje/Sve prstenovane ptice.xlsm', read_only=True)
    sheet = wb.active

    # Get total rows for progress tracking
    total_rows = sheet.max_row - 1  # Subtract header row
    print(f"📊 Found {total_rows:,} records to import")

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Clear existing data
    cursor.execute("DELETE FROM bird_ringing")

    # Import data
    imported_count = 0
    batch_size = 1000
    batch_data = []

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        # Extract only the first 38 meaningful columns (rest are empty "Column1", "Column2", etc.)
        if len(row) >= 38:
            cleaned_row = tuple(clean_value(val) for val in row[:38])
            batch_data.append(cleaned_row)

            # Insert batch when size reached
            if len(batch_data) >= batch_size:
                cursor.executemany("""
                    INSERT INTO bird_ringing (
                        ring_number, color_ring, species, age, sex, location, coordinates,
                        coordinate_accuracy, date, time, metal_ring_position, plastic_ring,
                        catching_method, bait, manipulation, status, clutch_size, pullus_age,
                        wing_length, third_primary_feather, mass, molting, back_claw,
                        bill_length, bill_measurement_method, head_length, tarsus,
                        tarsus_measurement_method, tail_length, fat_deposits, pectoral_muscle,
                        incubation_patch, alula, carpal_feathers, sex_determination,
                        protected_areas, ringer, notes
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, batch_data)
                conn.commit()
                imported_count += len(batch_data)
                batch_data = []

                # Progress update
                progress = (imported_count / total_rows) * 100
                print(f"  Progress: {imported_count:,}/{total_rows:,} ({progress:.1f}%)", end='\r')

    # Insert remaining batch
    if batch_data:
        cursor.executemany("""
            INSERT INTO bird_ringing (
                ring_number, color_ring, species, age, sex, location, coordinates,
                coordinate_accuracy, date, time, metal_ring_position, plastic_ring,
                catching_method, bait, manipulation, status, clutch_size, pullus_age,
                wing_length, third_primary_feather, mass, molting, back_claw,
                bill_length, bill_measurement_method, head_length, tarsus,
                tarsus_measurement_method, tail_length, fat_deposits, pectoral_muscle,
                incubation_patch, alula, carpal_feathers, sex_determination,
                protected_areas, ringer, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, batch_data)
        conn.commit()
        imported_count += len(batch_data)

    print(f"\n✓ Imported {imported_count:,} bird ringing records")

    # Get statistics
    cursor.execute("SELECT COUNT(DISTINCT species) FROM bird_ringing WHERE species IS NOT NULL")
    species_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(DISTINCT location) FROM bird_ringing WHERE location IS NOT NULL")
    location_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(DISTINCT ringer) FROM bird_ringing WHERE ringer IS NOT NULL")
    ringer_count = cursor.fetchone()[0]

    print(f"\n📊 Database Statistics:")
    print(f"  Total records: {imported_count:,}")
    print(f"  Unique species: {species_count:,}")
    print(f"  Unique locations: {location_count:,}")
    print(f"  Unique ringers: {ringer_count:,}")

    conn.close()
    wb.close()

def main():
    """Main import process."""
    print("=" * 60)
    print("Bird Ringing Database Import")
    print("Природњачки музеј у Београду")
    print("=" * 60)
    print()

    create_database()
    import_data()

    print()
    print("=" * 60)
    print("✓ Import complete!")
    print(f"📁 Database location: {DB_PATH}")
    print("=" * 60)

if __name__ == '__main__':
    main()
