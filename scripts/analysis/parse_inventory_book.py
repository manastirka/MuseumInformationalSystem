#!/usr/bin/env python3
"""
Parser for Knjiga Inventara (Inventory Book)
Extracts all inventory data from the physical inventory book Excel file
"""
import openpyxl
import sqlite3
import json
from datetime import datetime

def parse_main_inventory(ws):
    """
    Parse the main inventory sheet (Sheet1).

    Columns:
    1. Inventory number (broj inventara)
    2. Mineral name (naziv minerala)
    3. Locality (nalazište/lokalitet)
    4. Quantity/pieces (količina/komadi)
    5. Acquisition info (način nabavke)
    6. Collector/donator (prikupljač/donator)
    7. Notes/exhibition (napomene/izložba)
    8-11. Additional info
    """
    items = []

    for row_idx in range(1, ws.max_row + 1):
        row = ws[row_idx]

        # Extract values
        inv_number = row[0].value
        name = row[1].value
        locality = row[2].value
        quantity = row[3].value
        acquisition = row[4].value
        collector = row[5].value
        notes = row[6].value

        # Skip empty rows
        if not inv_number and not name:
            continue

        # Parse inventory number
        try:
            inv_num = int(inv_number) if inv_number else None
        except (ValueError, TypeError):
            inv_num = None

        item = {
            'inventory_number': inv_num,
            'inventory_number_raw': str(inv_number) if inv_number else '',
            'name': str(name).strip() if name else '',
            'locality': str(locality).strip() if locality else '',
            'quantity': str(quantity).strip() if quantity else '',
            'acquisition_info': str(acquisition).strip() if acquisition else '',
            'collector_donator': str(collector).strip() if collector else '',
            'notes': str(notes).strip() if notes else '',
            'sheet': 'Main Inventory',
            'row_number': row_idx
        }

        items.append(item)

    return items

def parse_meteorite_inventory(ws):
    """Parse meteorite inventory sheet."""
    items = []

    # First row is headers
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]

        inv_number = row[0].value
        name = row[1].value
        quantity = row[2].value if len(row) > 2 else None
        locality = row[3].value if len(row) > 3 else None
        formation = row[4].value if len(row) > 4 else None
        notes = row[5].value if len(row) > 5 else None

        if not inv_number and not name:
            continue

        try:
            inv_num = int(inv_number) if inv_number else None
        except (ValueError, TypeError):
            inv_num = None

        item = {
            'inventory_number': inv_num,
            'inventory_number_raw': str(inv_number) if inv_number else '',
            'name': str(name).strip() if name else '',
            'locality': str(locality).strip() if locality else '',
            'quantity': str(quantity).strip() if quantity else '',
            'acquisition_info': str(formation).strip() if formation else '',
            'collector_donator': '',
            'notes': str(notes).strip() if notes else '',
            'sheet': 'Meteoriti',
            'row_number': row_idx,
            'category': 'Meteorit'
        }

        items.append(item)

    return items

def parse_special_list(ws, sheet_name, category):
    """Parse special lists (asbestos, radioactive)."""
    items = []

    # Skip header row if present
    start_row = 2 if 'Inv' in str(ws[1][0].value or '') else 1

    for row_idx in range(start_row, ws.max_row + 1):
        row = ws[row_idx]

        inv_number = row[0].value
        name = row[1].value
        locality = row[2].value if len(row) > 2 else None
        acquisition = row[3].value if len(row) > 3 else None
        collector = row[4].value if len(row) > 4 else None
        notes = row[5].value if len(row) > 5 else None

        if not inv_number and not name:
            continue

        try:
            inv_num = int(inv_number) if inv_number else None
        except (ValueError, TypeError):
            inv_num = None

        item = {
            'inventory_number': inv_num,
            'inventory_number_raw': str(inv_number) if inv_number else '',
            'name': str(name).strip() if name else '',
            'locality': str(locality).strip() if locality else '',
            'quantity': '',
            'acquisition_info': str(acquisition).strip() if acquisition else '',
            'collector_donator': str(collector).strip() if collector else '',
            'notes': str(notes).strip() if notes else '',
            'sheet': sheet_name,
            'row_number': row_idx,
            'category': category
        }

        items.append(item)

    return items

def parse_inventory_book(filename):
    """Parse entire inventory book."""
    print(f"Parsing inventory book: {filename}")
    print("="*80)

    wb = openpyxl.load_workbook(filename, data_only=True)

    all_items = []

    # Parse main inventory
    print("\n1. Parsing main inventory (Sheet1)...")
    ws_main = wb['Sheet1']
    main_items = parse_main_inventory(ws_main)
    print(f"   ✓ Extracted {len(main_items)} items from main inventory")
    all_items.extend(main_items)

    # Parse meteorites
    print("\n2. Parsing meteorite inventory...")
    ws_meteorites = wb['Meteoriti']
    meteorite_items = parse_meteorite_inventory(ws_meteorites)
    print(f"   ✓ Extracted {len(meteorite_items)} meteorite items")
    all_items.extend(meteorite_items)

    # Parse asbestos list
    print("\n3. Parsing asbestos list...")
    ws_asbestos = wb['Lista AZBESTA']
    asbestos_items = parse_special_list(ws_asbestos, 'Lista AZBESTA', 'Azbest')
    print(f"   ✓ Extracted {len(asbestos_items)} asbestos items")
    all_items.extend(asbestos_items)

    # Parse radioactive list
    print("\n4. Parsing radioactive minerals list...")
    ws_radio = wb['Lista Radioaktivnih']
    radio_items = parse_special_list(ws_radio, 'Lista Radioaktivnih', 'Radioaktivni')
    print(f"   ✓ Extracted {len(radio_items)} radioactive items")
    all_items.extend(radio_items)

    print(f"\n" + "="*80)
    print(f"Total items extracted: {len(all_items)}")

    return all_items

def create_inventory_database(items, db_path='data/inventory_book.db'):
    """Create SQLite database for inventory book data."""
    print(f"\nCreating inventory database: {db_path}")

    import os
    os.makedirs(os.path.dirname(db_path), exist_ok=True)

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Create table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS inventory_book (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            inventory_number INTEGER,
            inventory_number_raw TEXT,
            name TEXT,
            locality TEXT,
            quantity TEXT,
            acquisition_info TEXT,
            collector_donator TEXT,
            notes TEXT,
            sheet TEXT,
            row_number INTEGER,
            category TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Insert data
    for item in items:
        cursor.execute('''
            INSERT INTO inventory_book (
                inventory_number, inventory_number_raw, name, locality,
                quantity, acquisition_info, collector_donator, notes,
                sheet, row_number, category
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            item.get('inventory_number'),
            item.get('inventory_number_raw', ''),
            item.get('name', ''),
            item.get('locality', ''),
            item.get('quantity', ''),
            item.get('acquisition_info', ''),
            item.get('collector_donator', ''),
            item.get('notes', ''),
            item.get('sheet', ''),
            item.get('row_number'),
            item.get('category', '')
        ))

    conn.commit()

    # Get statistics
    cursor.execute('SELECT COUNT(*) FROM inventory_book')
    total = cursor.fetchone()[0]

    cursor.execute('SELECT COUNT(DISTINCT inventory_number) FROM inventory_book WHERE inventory_number IS NOT NULL')
    unique_inv = cursor.fetchone()[0]

    cursor.execute('SELECT sheet, COUNT(*) FROM inventory_book GROUP BY sheet')
    by_sheet = dict(cursor.fetchall())

    conn.close()

    print(f"✓ Database created successfully")
    print(f"  Total records: {total}")
    print(f"  Unique inventory numbers: {unique_inv}")
    print(f"  Breakdown by sheet:")
    for sheet, count in by_sheet.items():
        print(f"    - {sheet}: {count}")

    return db_path

def save_to_json(items, json_path='data/inventory_book.json'):
    """Save items to JSON for web interface."""
    print(f"\nSaving to JSON: {json_path}")

    import os
    os.makedirs(os.path.dirname(json_path), exist_ok=True)

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump({
            'items': items,
            'total_count': len(items),
            'extracted_at': datetime.now().isoformat(),
            'source_file': 'Knjiga Inventara, P. Muzej - Copy.xlsx'
        }, f, ensure_ascii=False, indent=2)

    print(f"✓ JSON saved successfully")

def display_sample(items, count=10):
    """Display sample items."""
    print(f"\n" + "="*80)
    print(f"Sample items (first {count}):")
    print("="*80)

    for i, item in enumerate(items[:count], 1):
        print(f"\n{i}. Inv. #{item.get('inventory_number', 'N/A')} - {item.get('name', 'Unknown')}")
        print(f"   Locality: {item.get('locality', 'N/A')}")
        print(f"   Sheet: {item.get('sheet', 'N/A')}")
        if item.get('acquisition_info'):
            print(f"   Acquisition: {item.get('acquisition_info')[:60]}...")

if __name__ == "__main__":
    # Parse inventory book
    items = parse_inventory_book('Knjiga Inventara, P. Muzej - Copy.xlsx')

    # Display sample
    display_sample(items, 10)

    # Create database
    db_path = create_inventory_database(items)

    # Save to JSON
    save_to_json(items)

    print(f"\n" + "="*80)
    print("✓ Inventory book parsing complete!")
    print(f"✓ Database: {db_path}")
    print(f"✓ JSON: data/inventory_book.json")
    print("="*80)
