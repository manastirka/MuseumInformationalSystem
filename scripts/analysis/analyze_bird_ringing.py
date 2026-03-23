#!/usr/bin/env python3
import sys
import openpyxl

# Read the Excel file
wb = openpyxl.load_workbook('Prstenovanje/Sve prstenovane ptice.xlsm', read_only=True)
sheet = wb.active

# Get headers
headers = []
for cell in sheet[1]:
    headers.append(cell.value)

print("Columns:", headers)
print("\nTotal rows:", sheet.max_row)
print("\nFirst 10 rows:")

# Print first 10 rows
for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=11, values_only=True), 2):
    print(f"Row {i}:", row)

wb.close()
