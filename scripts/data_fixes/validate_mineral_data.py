#!/usr/bin/env python3
"""
Mineral Database Validation Script
Cross-checks mineral data from Excel file against PostgreSQL database
"""

import os
import sys
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.pool import NullPool
from datetime import datetime
import json

# Database URL
DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql+psycopg://aleksandarlukovic@localhost:5432/museum_system')

def load_excel_data(filepath):
    """Load mineral data from Excel file."""
    print(f"Loading Excel file: {filepath}")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb.active

    # Get headers from first row
    headers = []
    for cell in sheet[1]:
        headers.append(str(cell.value).strip().lower() if cell.value else '')

    print(f"Excel headers: {headers}")

    # Find column indices
    inv_col = None
    mineral_col = None
    kutija_col = None
    napomena_col = None

    for i, h in enumerate(headers):
        if 'inv' in h and 'broj' in h:
            inv_col = i
        elif 'mineral' in h:
            mineral_col = i
        elif 'kutija' in h:
            kutija_col = i
        elif 'napomena' in h:
            napomena_col = i

    print(f"Column indices - inv: {inv_col}, mineral: {mineral_col}, kutija: {kutija_col}")

    # Load data rows
    data = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[inv_col] is not None:
            inv_broj = str(row[inv_col]).strip() if row[inv_col] else ''
            mineral = str(row[mineral_col]).strip() if mineral_col is not None and row[mineral_col] else ''
            kutija = str(row[kutija_col]).strip() if kutija_col is not None and row[kutija_col] else ''
            napomena = str(row[napomena_col]).strip() if napomena_col is not None and row[napomena_col] else ''

            # Parse kutija to get storage location number
            kutija_num = None
            if kutija:
                try:
                    # Extract first number from kutija
                    import re
                    match = re.search(r'(\d+)', kutija)
                    if match:
                        kutija_num = int(match.group(1))
                except:
                    pass

            data.append({
                'row': row_num,
                'inv_broj': inv_broj,
                'mineral': mineral,
                'kutija': kutija,
                'kutija_num': kutija_num,
                'napomena': napomena
            })

    wb.close()
    print(f"Loaded {len(data)} records from Excel")
    return data


def load_database_data(engine):
    """Load all mineral data from PostgreSQL database."""
    print("Loading data from PostgreSQL database...")
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT
                id,
                inventory_number,
                item_name,
                storage_location,
                card_locality
            FROM minerals
            ORDER BY
                CASE WHEN inventory_number ~ '^[0-9]+$' THEN 0 ELSE 1 END,
                CASE WHEN inventory_number ~ '^[0-9]+$' THEN CAST(inventory_number AS INTEGER) ELSE 0 END,
                inventory_number
        """))

        data = {}
        for row in result:
            inv_num = str(row[1]).strip() if row[1] else ''
            # Remove 'M' prefix if present for comparison
            inv_num_clean = inv_num.lstrip('M').strip() if inv_num else ''

            data[inv_num_clean] = {
                'id': row[0],
                'inv_broj': inv_num,
                'naziv': str(row[2]).strip() if row[2] else '',
                'gde_se_nalazi': str(row[3]).strip() if row[3] else '',
                'lokalitet': str(row[4]).strip() if row[4] else ''
            }

        print(f"Loaded {len(data)} records from database")
        return data


def extract_storage_number(storage_str):
    """Extract first number from storage location string."""
    if not storage_str:
        return None
    import re
    match = re.search(r'(\d+)', str(storage_str))
    if match:
        return int(match.group(1))
    return None


def normalize_mineral_name(name):
    """Normalize mineral name for comparison."""
    if not name:
        return ''
    # Convert to lowercase, remove extra spaces
    name = ' '.join(name.lower().split())

    # Serbian spelling normalizations
    name = name.replace('ij', 'i')  # montmorijonit -> montmorionit
    name = name.replace('io', 'i')  # montmorionit -> montmornit (partial)

    # Remove diacritics for comparison
    diacritics = {'ž': 'z', 'č': 'c', 'ć': 'c', 'š': 's', 'đ': 'dj'}
    for d, r in diacritics.items():
        name = name.replace(d, r)

    # Normalize punctuation and spacing
    name = name.replace('(', ' ').replace(')', ' ')
    name = name.replace(',', ' ').replace('.', ' ')
    name = name.replace('-', ' ').replace('/', ' ')
    name = name.replace('  ', ' ').strip()

    # Common word variations
    name = name.replace('kristali', 'kristal')
    name = name.replace('druza', 'druz')
    name = name.replace('prevuceni', 'prevucen')
    name = name.replace('prevučeni', 'prevucen')
    name = name.replace('prevučen', 'prevucen')

    return name


def compare_names(excel_name, db_name):
    """Compare mineral names with fuzzy matching."""
    excel_norm = normalize_mineral_name(excel_name)
    db_norm = normalize_mineral_name(db_name)

    if excel_norm == db_norm:
        return True, "exact"

    # Check if one contains the other
    if excel_norm in db_norm or db_norm in excel_norm:
        return True, "partial"

    # Check first word match (main mineral name)
    excel_first = excel_norm.split()[0] if excel_norm else ''
    db_first = db_norm.split()[0] if db_norm else ''
    if excel_first and db_first and excel_first == db_first:
        return True, "first_word"

    # Check if main minerals match (extract mineral names, ignore modifiers)
    excel_minerals = set(excel_norm.split())
    db_minerals = set(db_norm.split())
    # Remove common non-mineral words
    stop_words = {'sa', 'i', 'u', 'na', 'po', 'delom', 'kristal'}
    excel_minerals -= stop_words
    db_minerals -= stop_words

    # If significant overlap in mineral names
    common = excel_minerals & db_minerals
    if len(common) >= 2 or (len(common) == 1 and len(excel_minerals) == 1):
        return True, "mineral_match"

    # Known equivalent names
    equivalents = [
        ('carolit', 'haroit'),
        ('hipšit', 'hidrogrosular'),
        ('hipsit', 'hidrogrosular'),
        ('giganstki', 'gigantski'),
        ('aguit', 'augit'),
        ('selestin', 'celestin'),
    ]
    for eq1, eq2 in equivalents:
        if (eq1 in excel_norm and eq2 in db_norm) or (eq2 in excel_norm and eq1 in db_norm):
            return True, "equivalent"

    return False, "mismatch"


def validate_data(excel_data, db_data):
    """Cross-check Excel data against database data."""
    print("\nStarting validation...")

    results = {
        'timestamp': datetime.now().isoformat(),
        'summary': {
            'excel_records': len(excel_data),
            'db_records': len(db_data),
            'matched': 0,
            'name_mismatch': 0,
            'location_mismatch': 0,
            'not_in_db': 0,
            'skipped_96_103': 0
        },
        'matches': [],
        'name_mismatches': [],
        'location_mismatches': [],
        'not_in_database': [],
        'skipped': []
    }

    for excel_row in excel_data:
        inv_broj = excel_row['inv_broj']
        kutija_num = excel_row['kutija_num']

        # Skip storage locations 96-103
        if kutija_num is not None and 96 <= kutija_num <= 103:
            results['skipped'].append({
                'inv_broj': inv_broj,
                'mineral': excel_row['mineral'],
                'kutija': excel_row['kutija'],
                'reason': f'Storage location {kutija_num} is in excluded range 96-103'
            })
            results['summary']['skipped_96_103'] += 1
            continue

        # Clean inventory number for lookup
        inv_clean = inv_broj.lstrip('M').strip() if inv_broj else ''

        if not inv_clean:
            continue

        # Look up in database
        if inv_clean in db_data:
            db_record = db_data[inv_clean]

            # Compare mineral names
            name_match, match_type = compare_names(excel_row['mineral'], db_record['naziv'])

            # Compare storage locations
            db_storage_num = extract_storage_number(db_record['gde_se_nalazi'])
            location_match = (kutija_num == db_storage_num)

            if name_match and location_match:
                results['summary']['matched'] += 1
                results['matches'].append({
                    'inv_broj': inv_broj,
                    'excel_mineral': excel_row['mineral'],
                    'db_mineral': db_record['naziv'],
                    'name_match_type': match_type,
                    'excel_kutija': excel_row['kutija'],
                    'db_location': db_record['gde_se_nalazi']
                })
            elif not name_match:
                results['summary']['name_mismatch'] += 1
                results['name_mismatches'].append({
                    'inv_broj': inv_broj,
                    'excel_mineral': excel_row['mineral'],
                    'db_mineral': db_record['naziv'],
                    'excel_kutija': excel_row['kutija'],
                    'db_location': db_record['gde_se_nalazi'],
                    'row': excel_row['row']
                })
            elif not location_match:
                results['summary']['location_mismatch'] += 1
                results['location_mismatches'].append({
                    'inv_broj': inv_broj,
                    'excel_mineral': excel_row['mineral'],
                    'db_mineral': db_record['naziv'],
                    'excel_kutija': excel_row['kutija'],
                    'excel_kutija_num': kutija_num,
                    'db_location': db_record['gde_se_nalazi'],
                    'db_location_num': db_storage_num,
                    'row': excel_row['row']
                })
        else:
            results['summary']['not_in_db'] += 1
            results['not_in_database'].append({
                'inv_broj': inv_broj,
                'mineral': excel_row['mineral'],
                'kutija': excel_row['kutija'],
                'row': excel_row['row']
            })

    return results


def print_report(results):
    """Print validation report."""
    print("\n" + "="*80)
    print("MINERAL DATABASE VALIDATION REPORT")
    print("="*80)
    print(f"Generated: {results['timestamp']}")
    print()

    s = results['summary']
    print("SUMMARY:")
    print(f"  Excel records processed: {s['excel_records']}")
    print(f"  Database records: {s['db_records']}")
    print(f"  Skipped (locations 96-103): {s['skipped_96_103']}")
    print()
    print(f"  ✓ Matched: {s['matched']}")
    print(f"  ✗ Name mismatches: {s['name_mismatch']}")
    print(f"  ✗ Location mismatches: {s['location_mismatch']}")
    print(f"  ✗ Not in database: {s['not_in_db']}")
    print()

    if results['name_mismatches']:
        print("\n" + "-"*80)
        print("NAME MISMATCHES (first 50):")
        print("-"*80)
        for item in results['name_mismatches'][:50]:
            print(f"  Row {item['row']}: Inv# {item['inv_broj']}")
            print(f"    Excel: {item['excel_mineral']}")
            print(f"    DB:    {item['db_mineral']}")
            print()

    if results['location_mismatches']:
        print("\n" + "-"*80)
        print("LOCATION MISMATCHES (first 50):")
        print("-"*80)
        for item in results['location_mismatches'][:50]:
            print(f"  Row {item['row']}: Inv# {item['inv_broj']} - {item['excel_mineral']}")
            print(f"    Excel kutija: {item['excel_kutija']} (num: {item['excel_kutija_num']})")
            print(f"    DB location:  {item['db_location']} (num: {item['db_location_num']})")
            print()

    if results['not_in_database']:
        print("\n" + "-"*80)
        print("NOT IN DATABASE (first 50):")
        print("-"*80)
        for item in results['not_in_database'][:50]:
            print(f"  Row {item['row']}: Inv# {item['inv_broj']} - {item['mineral']} (kutija: {item['kutija']})")

    print("\n" + "="*80)


def main():
    # Excel file path
    excel_file = "/home/aleksandarlukovic/MuseumInfoSystem/2026.01.20. kutije od 1-150.xlsx"

    if not os.path.exists(excel_file):
        print(f"Error: Excel file not found: {excel_file}")
        sys.exit(1)

    # Connect to database
    try:
        engine = create_engine(DATABASE_URL, poolclass=NullPool)
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        print("Database connection successful")
    except Exception as e:
        print(f"Error connecting to database: {e}")
        sys.exit(1)

    # Load data
    excel_data = load_excel_data(excel_file)
    db_data = load_database_data(engine)

    # Validate
    results = validate_data(excel_data, db_data)

    # Print report
    print_report(results)

    # Save full results to JSON
    output_file = "/home/aleksandarlukovic/MuseumInfoSystem/mineral_validation_report.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\nFull results saved to: {output_file}")


if __name__ == '__main__':
    main()
