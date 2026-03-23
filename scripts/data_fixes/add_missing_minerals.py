#!/usr/bin/env python3
"""
Add missing minerals from Excel to database
Only adds items with valid inventory numbers
"""

import os
import re
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.pool import NullPool
from datetime import datetime

DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql+psycopg://aleksandarlukovic@localhost:5432/museum_system')


def load_excel_data(filepath):
    """Load mineral data from Excel file."""
    print(f"Loading Excel file: {filepath}")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb.active

    # Get headers
    headers = []
    for cell in sheet[1]:
        headers.append(str(cell.value).strip().lower() if cell.value else '')

    # Find column indices
    inv_col = None
    mineral_col = None
    kutija_col = None
    for i, h in enumerate(headers):
        if 'inv' in h and 'broj' in h:
            inv_col = i
        elif 'mineral' in h:
            mineral_col = i
        elif 'kutija' in h:
            kutija_col = i

    # Load data
    data = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[inv_col] is not None:
            inv_broj = str(row[inv_col]).strip()
            mineral = str(row[mineral_col]).strip() if mineral_col is not None and row[mineral_col] else ''
            kutija = str(row[kutija_col]).strip() if kutija_col is not None and row[kutija_col] else ''

            data.append({
                'row': row_num,
                'inv_broj': inv_broj,
                'mineral': mineral,
                'kutija': kutija
            })

    wb.close()
    return data


def is_valid_inv_number(inv_str):
    """Check if inventory number is valid (numeric or starts with number)."""
    if not inv_str:
        return False, None

    # Skip entries marked as "bez broja", "neidentifikovan", etc.
    lower_inv = inv_str.lower()
    skip_patterns = ['bez broja', 'bez etikete', 'bez kartice', 'neidentifikovan', 'neinventarisan']
    for pattern in skip_patterns:
        if pattern in lower_inv:
            return False, None

    # Extract numeric part - handle cases like "2932, 317" by taking first number
    match = re.match(r'^(\d+)', inv_str.strip())
    if match:
        return True, match.group(1)

    return False, None


def add_missing_minerals(engine, excel_data):
    """Add missing minerals to database."""
    print("\n--- Finding and adding missing minerals ---")

    # First, get all existing inventory numbers from database
    with engine.connect() as conn:
        result = conn.execute(text("SELECT inventory_number FROM minerals"))
        existing_inv = set(str(row[0]).strip() for row in result if row[0])

    print(f"Found {len(existing_inv)} existing inventory numbers in database")

    # Find items to add
    items_to_add = []
    for item in excel_data:
        valid, inv_num = is_valid_inv_number(item['inv_broj'])
        if valid and inv_num not in existing_inv:
            items_to_add.append({
                'inv_broj': inv_num,
                'mineral': item['mineral'],
                'kutija': item['kutija'],
                'row': item['row']
            })

    print(f"Found {len(items_to_add)} items to add")

    if not items_to_add:
        print("No items to add")
        return 0

    # Add items to database
    added = 0
    with engine.connect() as conn:
        for item in items_to_add:
            try:
                # Check one more time if it doesn't exist
                result = conn.execute(text(
                    "SELECT id FROM minerals WHERE inventory_number = :inv"
                ), {'inv': item['inv_broj']})

                if result.fetchone():
                    print(f"  Skipping Inv# {item['inv_broj']} - already exists")
                    continue

                # Insert new record
                conn.execute(text("""
                    INSERT INTO minerals (inventory_number, item_name, storage_location, created_at)
                    VALUES (:inv, :name, :location, :created)
                """), {
                    'inv': item['inv_broj'],
                    'name': item['mineral'],
                    'location': item['kutija'],
                    'created': datetime.now()
                })

                print(f"  Added: Inv# {item['inv_broj']} - {item['mineral']} (kutija: {item['kutija']})")
                added += 1

            except Exception as e:
                print(f"  Error adding Inv# {item['inv_broj']}: {e}")

        conn.commit()

    return added


def main():
    excel_file = "/home/aleksandarlukovic/MuseumInfoSystem/2026.01.20. kutije od 1-150.xlsx"

    if not os.path.exists(excel_file):
        print(f"Error: Excel file not found: {excel_file}")
        return

    # Connect to database
    try:
        engine = create_engine(DATABASE_URL, poolclass=NullPool)
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        print("Database connection successful")
    except Exception as e:
        print(f"Error connecting to database: {e}")
        return

    # Load Excel data
    excel_data = load_excel_data(excel_file)
    print(f"Loaded {len(excel_data)} records from Excel")

    # Add missing minerals
    added = add_missing_minerals(engine, excel_data)

    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"  Items added to database: {added}")
    print("="*60)


if __name__ == '__main__':
    main()
