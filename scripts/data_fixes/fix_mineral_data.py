#!/usr/bin/env python3
"""
Fix Mineral Database - Correct typos and update locations from Excel
"""

import os
import re
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.pool import NullPool
from datetime import datetime

DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql+psycopg://aleksandarlukovic@localhost:5432/museum_system')

# Typo corrections: DB value -> Correct value
NAME_CORRECTIONS = {
    # Typos found in validation
    'beril (akvmarin)': 'beril (akvamarin)',
    'kalci t, halkopirit i karbonat': 'kalcit, halkopirit i karbonat',
    'zad': 'žad',
    'caroit': 'haroit',  # Charoite in Serbian is haroit
    's': None,  # Invalid entry - needs manual review
    'montmorionit': 'montmorijonit',
    # Normalize spacing
    'kvarc, kalcit, pirit, sfalerit': 'kvarc, kalcit, pirit, sfalerit',
    'pirit, arsenopirit, kvarc i dolomit': 'pirit, arsenopirit, kvarc i dolomit',
    'kristali kvarca prvučeni kalcitom': 'kristali kvarca prevučeni kalcitom',
}

def load_excel_locations(filepath):
    """Load storage locations from Excel file."""
    print(f"Loading Excel file: {filepath}")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb.active

    # Get headers
    headers = []
    for cell in sheet[1]:
        headers.append(str(cell.value).strip().lower() if cell.value else '')

    # Find column indices
    inv_col = None
    kutija_col = None
    for i, h in enumerate(headers):
        if 'inv' in h and 'broj' in h:
            inv_col = i
        elif 'kutija' in h:
            kutija_col = i

    # Load data - map inventory number to kutija
    locations = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[inv_col] is not None:
            inv_broj = str(row[inv_col]).strip()
            kutija = str(row[kutija_col]).strip() if kutija_col is not None and row[kutija_col] else ''

            # Clean inventory number
            inv_clean = inv_broj.lstrip('M').strip()

            # Skip invalid entries
            if not inv_clean or not inv_clean.isdigit():
                continue

            # Extract kutija number to check if in 96-103 range
            kutija_num = None
            if kutija:
                match = re.search(r'(\d+)', kutija)
                if match:
                    kutija_num = int(match.group(1))

            # Skip locations 96-103
            if kutija_num is not None and 96 <= kutija_num <= 103:
                continue

            if kutija:
                locations[inv_clean] = kutija

    wb.close()
    print(f"Loaded {len(locations)} location mappings from Excel")
    return locations


def fix_name_typos(engine):
    """Fix known typos in mineral names."""
    print("\n--- Fixing Name Typos ---")
    fixes_applied = 0

    with engine.connect() as conn:
        for wrong, correct in NAME_CORRECTIONS.items():
            if correct is None:
                print(f"  Skipping '{wrong}' - needs manual review")
                continue

            # Check if record exists
            result = conn.execute(text(
                "SELECT id, inventory_number, item_name FROM minerals WHERE item_name = :wrong"
            ), {'wrong': wrong})

            rows = result.fetchall()
            if rows:
                for row in rows:
                    print(f"  Fixing: Inv# {row[1]} - '{wrong}' -> '{correct}'")

                # Update
                conn.execute(text(
                    "UPDATE minerals SET item_name = :correct WHERE item_name = :wrong"
                ), {'wrong': wrong, 'correct': correct})
                fixes_applied += len(rows)

        conn.commit()

    print(f"Applied {fixes_applied} name corrections")
    return fixes_applied


def update_locations(engine, excel_locations):
    """Update storage locations from Excel data."""
    print("\n--- Updating Storage Locations ---")
    updates_applied = 0

    with engine.connect() as conn:
        for inv_clean, kutija in excel_locations.items():
            # Get current DB record
            result = conn.execute(text("""
                SELECT id, inventory_number, storage_location
                FROM minerals
                WHERE inventory_number = :inv_num
            """), {'inv_num': inv_clean})

            row = result.fetchone()
            if row:
                db_location = str(row[2]).strip() if row[2] else ''

                # Only update if different and Excel has a value
                if kutija and kutija != db_location:
                    # Extract first number from both to compare
                    excel_num = None
                    db_num = None

                    match = re.search(r'(\d+)', kutija)
                    if match:
                        excel_num = int(match.group(1))

                    if db_location:
                        match = re.search(r'(\d+)', db_location)
                        if match:
                            db_num = int(match.group(1))

                    # Update if DB is empty or numbers differ
                    if not db_location or (excel_num != db_num):
                        print(f"  Updating Inv# {inv_clean}: '{db_location}' -> '{kutija}'")
                        conn.execute(text("""
                            UPDATE minerals
                            SET storage_location = :location
                            WHERE inventory_number = :inv_num
                        """), {'location': kutija, 'inv_num': inv_clean})
                        updates_applied += 1

        conn.commit()

    print(f"Applied {updates_applied} location updates")
    return updates_applied


def main():
    excel_file = "/home/aleksandarlukovic/MuseumInfoSystem/2026.01.20. kutije od 1-150.xlsx"

    if not os.path.exists(excel_file):
        print(f"Error: Excel file not found: {excel_file}")
        return

    # Connect to database
    try:
        engine = create_engine(DATABASE_URL, poolclass=NullPool)
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        print("Database connection successful")
    except Exception as e:
        print(f"Error connecting to database: {e}")
        return

    # Load Excel locations
    excel_locations = load_excel_locations(excel_file)

    # Fix typos
    typos_fixed = fix_name_typos(engine)

    # Update locations
    locations_updated = update_locations(engine, excel_locations)

    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"  Name typos fixed: {typos_fixed}")
    print(f"  Locations updated: {locations_updated}")
    print(f"  Total changes: {typos_fixed + locations_updated}")
    print("="*60)


if __name__ == '__main__':
    main()
