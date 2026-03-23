#!/usr/bin/env python3
"""
Fill missing mineral names from Excel where database has N/A or empty
"""

import os
import re
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.pool import NullPool

DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql+psycopg://aleksandarlukovic@localhost:5432/museum_system')


def load_excel_names(filepath):
    """Load mineral names from Excel file."""
    print(f"Loading Excel file: {filepath}")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb.active

    # Get headers
    headers = []
    for cell in sheet[1]:
        headers.append(str(cell.value).strip().lower() if cell.value else '')

    # Find column indices
    inv_col = None
    mineral_col = None
    for i, h in enumerate(headers):
        if 'inv' in h and 'broj' in h:
            inv_col = i
        elif 'mineral' in h:
            mineral_col = i

    # Load data - map inventory number to mineral name
    names = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[inv_col] is not None:
            inv_broj = str(row[inv_col]).strip()
            mineral = str(row[mineral_col]).strip() if mineral_col is not None and row[mineral_col] else ''

            # Clean inventory number
            inv_clean = inv_broj.lstrip('M').strip()

            # Skip invalid entries
            if not inv_clean or not inv_clean.isdigit():
                continue

            if mineral and mineral.lower() not in ['neidentifikovan', 'neinventarisan', 'bez broja', '']:
                names[inv_clean] = mineral

    wb.close()
    print(f"Loaded {len(names)} mineral names from Excel")
    return names


def find_and_fill_missing(engine, excel_names):
    """Find records with N/A or empty names and fill from Excel."""
    print("\n--- Finding records with missing names ---")

    updates = []

    with engine.connect() as conn:
        # Find records with N/A, empty, or very short names
        result = conn.execute(text("""
            SELECT id, inventory_number, item_name
            FROM minerals
            WHERE item_name IS NULL
               OR item_name = ''
               OR item_name = 'N/A'
               OR item_name = 'n/a'
               OR item_name = '-'
               OR item_name = 's'
               OR LENGTH(TRIM(item_name)) <= 2
            ORDER BY CAST(inventory_number AS INTEGER)
        """))

        missing_records = result.fetchall()
        print(f"Found {len(missing_records)} records with missing/invalid names in database")

        for row in missing_records:
            db_id = row[0]
            inv_num = str(row[1]).strip() if row[1] else ''
            current_name = str(row[2]).strip() if row[2] else ''

            inv_clean = inv_num.lstrip('M').strip()

            if inv_clean in excel_names:
                excel_name = excel_names[inv_clean]
                updates.append({
                    'id': db_id,
                    'inv_num': inv_num,
                    'old_name': current_name,
                    'new_name': excel_name
                })

        print(f"Found {len(updates)} records that can be updated from Excel")

        # Apply updates
        if updates:
            print("\n--- Applying updates ---")
            for update in updates:
                print(f"  Inv# {update['inv_num']}: '{update['old_name']}' -> '{update['new_name']}'")
                conn.execute(text("""
                    UPDATE minerals
                    SET item_name = :new_name
                    WHERE id = :id
                """), {'new_name': update['new_name'], 'id': update['id']})

            conn.commit()
            print(f"\nApplied {len(updates)} name updates")

    return updates


def main():
    excel_file = "/home/aleksandarlukovic/MuseumInfoSystem/2026.01.20. kutije od 1-150.xlsx"

    if not os.path.exists(excel_file):
        print(f"Error: Excel file not found: {excel_file}")
        return

    # Connect to database
    try:
        engine = create_engine(DATABASE_URL, poolclass=NullPool)
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        print("Database connection successful")
    except Exception as e:
        print(f"Error connecting to database: {e}")
        return

    # Load Excel names
    excel_names = load_excel_names(excel_file)

    # Find and fill missing names
    updates = find_and_fill_missing(engine, excel_names)

    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"  Names filled from Excel: {len(updates)}")
    print("="*60)


if __name__ == '__main__':
    main()
