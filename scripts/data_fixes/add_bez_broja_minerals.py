#!/usr/bin/env python3
"""
Add minerals without inventory numbers (bez broja) to database
Inventory numbers will be added later
"""

import os
import re
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.pool import NullPool
from datetime import datetime

DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql+psycopg://aleksandarlukovic@localhost:5432/museum_system')


def load_excel_data(filepath):
    """Load mineral data from Excel file."""
    print(f"Loading Excel file: {filepath}")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb.active

    # Get headers
    headers = []
    for cell in sheet[1]:
        headers.append(str(cell.value).strip().lower() if cell.value else '')

    # Find column indices
    inv_col = None
    mineral_col = None
    kutija_col = None
    napomena_col = None
    for i, h in enumerate(headers):
        if 'inv' in h and 'broj' in h:
            inv_col = i
        elif 'mineral' in h:
            mineral_col = i
        elif 'kutija' in h:
            kutija_col = i
        elif 'napomena' in h:
            napomena_col = i

    # Load data
    data = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[inv_col] is not None:
            inv_broj = str(row[inv_col]).strip()
            mineral = str(row[mineral_col]).strip() if mineral_col is not None and row[mineral_col] else ''
            kutija = str(row[kutija_col]).strip() if kutija_col is not None and row[kutija_col] else ''
            napomena = str(row[napomena_col]).strip() if napomena_col is not None and row[napomena_col] else ''

            data.append({
                'row': row_num,
                'inv_broj': inv_broj,
                'mineral': mineral,
                'kutija': kutija,
                'napomena': napomena
            })

    wb.close()
    return data


def is_bez_broja(inv_str):
    """Check if inventory number indicates no number assigned."""
    if not inv_str:
        return False

    lower_inv = inv_str.lower()
    patterns = ['bez broja', 'bez etikete', 'bez kartice', 'neidentifikovan', 'neinventarisan']
    for pattern in patterns:
        if pattern in lower_inv:
            return True

    return False


def add_bez_broja_minerals(engine, excel_data):
    """Add minerals without inventory numbers to database."""
    print("\n--- Adding minerals without inventory numbers ---")

    # Find items without inventory numbers
    items_to_add = []
    for item in excel_data:
        if is_bez_broja(item['inv_broj']):
            # Skip if mineral name is also empty or generic
            if item['mineral'] and item['mineral'].lower() not in ['neidentifikovan', 'neinventarisan', 'razni', '']:
                items_to_add.append(item)

    print(f"Found {len(items_to_add)} items without inventory numbers to add")

    if not items_to_add:
        print("No items to add")
        return 0

    # Get next temporary ID
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT MAX(CAST(SUBSTRING(inventory_number FROM 'BEZ-([0-9]+)') AS INTEGER))
            FROM minerals
            WHERE inventory_number LIKE 'BEZ-%'
        """))
        max_bez = result.scalar() or 0
        next_bez_id = max_bez + 1

    # Add items to database
    added = 0
    with engine.connect() as conn:
        for item in items_to_add:
            try:
                # Create temporary inventory number
                temp_inv = f"BEZ-{next_bez_id:04d}"

                # Build notes from original inv field and napomena
                notes = f"Original: {item['inv_broj']}"
                if item['napomena']:
                    notes += f" | {item['napomena']}"

                # Insert new record
                conn.execute(text("""
                    INSERT INTO minerals (inventory_number, item_name, storage_location, comments, created_at)
                    VALUES (:inv, :name, :location, :comments, :created)
                """), {
                    'inv': temp_inv,
                    'name': item['mineral'],
                    'location': item['kutija'],
                    'comments': notes,
                    'created': datetime.now()
                })

                print(f"  Added: {temp_inv} - {item['mineral']} (kutija: {item['kutija']})")
                added += 1
                next_bez_id += 1

            except Exception as e:
                print(f"  Error adding {item['mineral']}: {e}")

        conn.commit()

    return added


def main():
    excel_file = "/home/aleksandarlukovic/MuseumInfoSystem/2026.01.20. kutije od 1-150.xlsx"

    if not os.path.exists(excel_file):
        print(f"Error: Excel file not found: {excel_file}")
        return

    # Connect to database
    try:
        engine = create_engine(DATABASE_URL, poolclass=NullPool)
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        print("Database connection successful")
    except Exception as e:
        print(f"Error connecting to database: {e}")
        return

    # Load Excel data
    excel_data = load_excel_data(excel_file)
    print(f"Loaded {len(excel_data)} records from Excel")

    # Add bez broja minerals
    added = add_bez_broja_minerals(engine, excel_data)

    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"  Items added with temporary IDs (BEZ-XXXX): {added}")
    print("  These items need inventory numbers assigned later")
    print("="*60)


if __name__ == '__main__':
    main()
