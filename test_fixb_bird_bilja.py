import os
import re
import shutil
import tempfile
import zipfile

os.environ.setdefault('FLASK_ENV', 'testing')
os.environ.setdefault('SECRET_KEY', 'test-secret')
os.environ.setdefault('REDIS_URL', '')
os.environ.setdefault('SESSION_TYPE', 'filesystem')
os.environ.setdefault('SESSION_FILE_DIR', '/tmp/museum-test-b-bird-bilja')

import openpyxl
from openpyxl import Workbook

import bird_ringing_database as brd
import import_bilja_collections as ibc


# ---------------------------------------------------------------------------
# Finding 1: read queries hardcoded NULL for color_ring
# ---------------------------------------------------------------------------

def test_base_query_does_not_select_absent_color_ring_column():
    """M08 follow-up: the deployed `bird_ringing_records` table does NOT have a
    color_ring column (schema drift vs db/schema.sql), so SELECTing br.color_ring
    raises psycopg UndefinedColumn and breaks every read. Until a migration adds
    the column, the read must keep returning NULL AS color_ring. Surfacing the
    real value is a schema-migration task, not a query-only change.
    """
    query = brd._PG_BASE_QUERY
    assert 'NULL AS color_ring' in query, 'read must not query the absent color_ring column'
    assert 'br.color_ring' not in query, 'br.color_ring crashes: column absent in deployed DB'


# ---------------------------------------------------------------------------
# Finding 2: xlsx read_only iter_rows truncates trailing empty cells,
# dropping whole rows on import
# ---------------------------------------------------------------------------

def _make_xlsx_with_short_trailing_row(header_width, filled, sheet_name='Sheet1'):
    """Build an .xlsx whose data row has fewer cells than the header.

    Removing the <dimension> element forces openpyxl read_only mode to
    derive width per-row, so the trailing-empty data row is yielded shorter
    than the header -- exactly the production truncation scenario.
    """
    tmp = tempfile.mktemp(suffix='.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(['c%d' % i for i in range(header_width)])
    ws.append(['v%d' % i for i in range(filled)] + [None] * (header_width - filled))
    wb.save(tmp)

    workdir = tempfile.mkdtemp()
    with zipfile.ZipFile(tmp) as z:
        z.extractall(workdir)
    sheet_xml = os.path.join(workdir, 'xl', 'worksheets', 'sheet1.xml')
    xml = open(sheet_xml, encoding='utf-8').read()
    xml = re.sub(r'<dimension[^/]*/>', '', xml)
    open(sheet_xml, 'w', encoding='utf-8').write(xml)

    out = tempfile.mktemp(suffix='.xlsx')
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(workdir):
            for f in files:
                full = os.path.join(root, f)
                z.write(full, os.path.relpath(full, workdir))

    os.remove(tmp)
    shutil.rmtree(workdir)
    return out


def test_short_trailing_row_is_demonstrated_by_raw_openpyxl():
    """Sanity check: raw openpyxl read_only really yields a short data row."""
    path = _make_xlsx_with_short_trailing_row(16, 13)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb['Sheet1']
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    finally:
        os.remove(path)
    assert len(rows[0]) == 16
    assert len(rows[1]) == 13  # truncated trailing empties


def test_iter_xlsx_rows_pads_short_rows_to_header_width():
    """_iter_xlsx_rows must not yield rows narrower than the header, otherwise
    fixed-position mappers raise IndexError and drop valid records."""
    path = _make_xlsx_with_short_trailing_row(16, 13)
    try:
        data_rows = list(ibc._iter_xlsx_rows(path, 'Sheet1'))
    finally:
        os.remove(path)

    assert len(data_rows) == 1
    row = data_rows[0]
    assert len(row) >= 16, 'short row was not padded to header width'
    # Trailing cells preserved as empty, not lost
    assert row[13] is None and row[14] is None and row[15] is None
    assert row[:13] == tuple('v%d' % i for i in range(13))


def test_mapper_does_not_drop_record_with_blank_trailing_fields():
    """End-to-end: a valid kenozojske row whose last column (broj_primeraka)
    is blank must map successfully rather than raise IndexError."""
    path = _make_xlsx_with_short_trailing_row(16, 13)
    try:
        data_rows = list(ibc._iter_xlsx_rows(path, 'Sheet1'))
    finally:
        os.remove(path)

    assert len(data_rows) == 1
    # _map_kenozojske indexes up to row[15]; must not raise.
    mapped = ibc._map_kenozojske(data_rows[0], 'src.xlsx')
    assert mapped[-1] == 'src.xlsx'
    assert mapped[15] is None  # broj_primeraka blank, preserved as NULL
