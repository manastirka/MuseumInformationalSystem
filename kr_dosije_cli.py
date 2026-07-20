"""`flask uvezi-kr-dosije <putanja>` — увоз конзерваторског Excel-а + слика.

Модел је стари дневник геолошког одељења: један xlsx (10 колона) и ~172 слике
чија имена носе број предмета и фазу (сл. а/б/в = pre/tokom/posle). Команда:

  flask uvezi-kr-dosije "/put/do/folder"                 # DRY-RUN (ништа не пише)
  flask uvezi-kr-dosije "/put/do/folder" --execute       # тражи куцање имена базе

Понашање:
  * DRY-RUN подразумеван — само извештај (колико досијеа, поклопљених извршилаца,
    парсираних периода, слика по стању: тачно/двосмислено/непоклопљено).
  * --execute уписује досијее + извршиоце и отпрема ТАЧНО поклопљене слике кроз
    Фototeka ток (аутор = конзератор одељења), везујући их са фазом. Двосмислене/
    непоклопљене слике се НЕ отпремају аутоматски — пријаве се за ручно повезивање.
  * Идемпотентно по запису: досије са истим (izvor='uvoz', изворни ев. број,
    назив) се не убацује поново.

Не мења шему, не брише ништа. Слике иду искључиво кроз постојећи Фototeka intake.
"""

import os
import re
import uuid
from datetime import date
from pathlib import Path
from urllib.parse import urlsplit

import click

from postgres_service import get_database_url, get_postgres_connection
import kr_dosije_core as core

# Подразумевани аутор отпремљених слика по одељењу (конзератор).
KONZERVATOR_EMAIL = {
    'geo': 'nenad.mladenovic@nhmbeo.rs',
    'bio': 'gorana.petkovski@nhmbeo.rs',
}
IMAGE_EXTS = ('.jpg', '.jpeg', '.png', '.tif', '.tiff', '.webp')


def _database_name():
    return urlsplit(get_database_url()).path.lstrip('/') or '?'


# ---------------------------------------------------------------------------
# Откривање улаза
# ---------------------------------------------------------------------------
def discover_inputs(putanja):
    """Врати (xlsx_path, [image_paths]) за дату путању (фолдер или .xlsx)."""
    p = Path(putanja).expanduser()
    if p.is_file() and p.suffix.lower() in ('.xlsx', '.xls'):
        folder = p.parent
        xlsx = p
    elif p.is_dir():
        folder = p
        xlsxs = sorted(f for f in folder.iterdir()
                       if f.suffix.lower() in ('.xlsx', '.xls'))
        if not xlsxs:
            raise click.ClickException(f'Нема .xlsx у {folder}')
        xlsx = xlsxs[0]
    else:
        raise click.ClickException(f'Путања не постоји: {putanja}')
    images = sorted(f for f in folder.iterdir()
                    if f.suffix.lower() in IMAGE_EXTS)
    return xlsx, images


# ---------------------------------------------------------------------------
# Рашчлањавање xlsx → досијеи
# ---------------------------------------------------------------------------
def parse_dossiers(xlsx_path, odeljenje, known_users):
    """Прочитај стварне записе (колона D непразна) у листу dict-ова.

    Колоне: A ев.бр, B инв, C колектор, D назив, E пре, F поступак, G после,
    H извршиоци, I период, J напомена.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))[1:]  # прескочи заглавље

    dossiers = []
    for r in rows:
        r = list(r) + [None] * (10 - len(r))
        a, b, c, d, e, f, g, h, i, j = r[:10]
        if not d or not str(d).strip():
            continue  # празни/шаблонски редови
        period_od, period_do = core.parse_period(i)
        inv_raw = _clean(b)
        col_raw = _clean(c)
        is_mineral = bool(inv_raw and re.match(r'^\s*[МM]?\s*\d', inv_raw))
        keys = set(k for k in (core.broj_key_from_text(b), core.broj_key_from_text(c)) if k)
        dossiers.append({
            'izvorni_evidencioni_broj': str(a).strip() if a is not None else None,
            'odeljenje': odeljenje,
            'predmet_tip': 'zbirka' if is_mineral else 'slobodan',
            'database_name': 'mineral' if is_mineral else None,
            'inventarni_broj': inv_raw,
            'kolektorski_broj': col_raw,
            'naziv_predmeta': _clean(d),
            'opis_pre': _clean(e),
            'opis_postupak': _clean(f),
            'opis_posle': _clean(g),
            'period_od': period_od,
            'period_do': period_do,
            'period_tekst': _clean(i),
            'napomena': _clean(j),
            'izvrsioci': core.parse_izvrsioci(h, known_users),
            'keys': keys,
        })
    return dossiers


def _clean(v):
    if v is None:
        return None
    s = ' '.join(str(v).split())
    return s or None


# ---------------------------------------------------------------------------
# Поклапање слика са досијеима
# ---------------------------------------------------------------------------
def match_images(images, dossiers):
    """Врати dict: exact=[(img, dosije_idx, faza)], те извештајне корпе
    ambiguous / unmatched / no_phase."""
    key_to_idx = {}
    for idx, d in enumerate(dossiers):
        for k in d['keys']:
            key_to_idx.setdefault(k, []).append(idx)

    exact, ambiguous, unmatched, no_phase = [], [], [], []
    for img in images:
        parsed = core.parse_image_filename(img.name)
        if parsed is None:
            no_phase.append(img.name)
            continue
        n = parsed['broj_norm']
        idxs = key_to_idx.get(n)
        if idxs and len(idxs) == 1:
            exact.append((img, idxs[0], parsed['faza']))
        elif idxs:
            ambiguous.append((img.name, n, [dossiers[i]['izvorni_evidencioni_broj']
                                            for i in idxs]))
        else:
            unmatched.append((img.name, n))
    return {'exact': exact, 'ambiguous': ambiguous,
            'unmatched': unmatched, 'no_phase': no_phase}


# ---------------------------------------------------------------------------
# План (parse + match), без иједне измене — за dry-run и тестове
# ---------------------------------------------------------------------------
def plan_import(putanja, odeljenje, known_users):
    xlsx, images = discover_inputs(putanja)
    dossiers = parse_dossiers(xlsx, odeljenje, known_users)
    matches = match_images(images, dossiers)
    return {'xlsx': xlsx, 'images': images, 'dossiers': dossiers, 'matches': matches}


# ---------------------------------------------------------------------------
# Упис
# ---------------------------------------------------------------------------
def _already_imported(cur, d):
    cur.execute(
        """
        SELECT id FROM kr_dosije
        WHERE izvor = 'uvoz' AND izvorni_evidencioni_broj IS NOT DISTINCT FROM %s
              AND naziv_predmeta = %s
        LIMIT 1
        """,
        (d['izvorni_evidencioni_broj'], d['naziv_predmeta']),
    )
    row = cur.fetchone()
    return row[0] if row else None


def _insert_dossier(cur, d, autor_email):
    god = d['period_od'].year if d['period_od'] else date.today().year
    evb = core.next_evidencioni_broj(cur, d['odeljenje'], god)
    cur.execute(
        """
        INSERT INTO kr_dosije
            (evidencioni_broj, odeljenje, predmet_tip, database_name,
             inventarni_broj, kolektorski_broj, naziv_predmeta,
             opis_pre, opis_postupak, opis_posle, period_od, period_do,
             period_tekst, napomena, izvor, izvorni_evidencioni_broj, created_by_email)
        VALUES (%(evb)s, %(odeljenje)s, %(predmet_tip)s, %(database_name)s,
                %(inventarni_broj)s, %(kolektorski_broj)s, %(naziv_predmeta)s,
                %(opis_pre)s, %(opis_postupak)s, %(opis_posle)s, %(period_od)s,
                %(period_do)s, %(period_tekst)s, %(napomena)s, 'uvoz',
                %(izvorni_evidencioni_broj)s, %(autor)s)
        RETURNING id
        """,
        {**d, 'evb': evb, 'autor': autor_email},
    )
    dosije_id = cur.fetchone()[0]
    for red, iz in enumerate(d['izvrsioci']):
        cur.execute(
            """
            INSERT INTO kr_dosije_izvrsilac (dosije_id, user_email, ime_tekst, redosled)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (dosije_id, ime_tekst) DO NOTHING
            """,
            (dosije_id, iz['user_email'], iz['ime_tekst'], red),
        )
    return dosije_id, evb


def _upload_and_link(cur, img_path, dosije_id, faza, autor_email):
    """Отпреми слику кроз Фototeka intake (или нађи постојећу по sha256) и вежи."""
    import fototeka_views
    import fototeka_jobs
    ext = img_path.suffix.lower()
    file_size = img_path.stat().st_size
    sha256 = fototeka_jobs.sha256_of_file(img_path)
    cur.execute('SELECT id FROM fotografije WHERE sha256 = %s', (sha256,))
    existing = cur.fetchone()
    if existing:
        foto_id = existing[0]
    else:
        temp_dir = fototeka_jobs.get_media_path() / 'temp'
        temp_dir.mkdir(parents=True, exist_ok=True)
        temp_path = temp_dir / f'{uuid.uuid4().hex}_{ext.lstrip(".")}'
        import shutil
        shutil.copyfile(img_path, temp_path)
        try:
            foto_id, reason = fototeka_views._intake_photo_from_path(
                cur, temp_path, img_path.name, file_size, ext,
                autor_email=autor_email, opis=None, tags=['к-р-досије', 'увоз'],
                datum_override=None, veza=None, u_prijemnom_redu=False,
                poreklo='import', vidljivost='javno',
            )
        finally:
            if temp_path.exists():
                temp_path.unlink()
        if foto_id is None:
            return False
    cur.execute(
        """
        INSERT INTO foto_veza_kr_dosije
            (fotografija_id, dosije_id, faza, redosled, created_by_email)
        VALUES (%s, %s, %s, %s, %s)
        ON CONFLICT (fotografija_id, dosije_id, faza) DO NOTHING
        """,
        (foto_id, dosije_id, faza, core.FAZA_REDOSLED.get(faza, 0), autor_email),
    )
    return True


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def _load_users():
    with get_postgres_connection() as conn, conn.cursor() as cur:
        cur.execute('SELECT email, full_name FROM users WHERE full_name IS NOT NULL')
        return [(e, n) for e, n in cur.fetchall()]


def register_cli(app):
    @app.cli.command('uvezi-kr-dosije')
    @click.argument('putanja')
    @click.option('--odeljenje', default='geo',
                  type=click.Choice(['geo', 'bio']),
                  help='Одељење за све записе (подразумевано geo).')
    @click.option('--bez-slika', is_flag=True,
                  help='Не отпремај слике — само досијее.')
    @click.option('--execute', is_flag=True,
                  help='Стварно упиши (подразумевано dry-run).')
    def uvezi_kr_dosije(putanja, odeljenje, bez_slika, execute):
        """Увоз К-Р досијеа из xlsx + слика (dry-run подразумеван)."""
        mode = 'EXECUTE' if execute else 'DRY-RUN (без измена)'
        click.echo(f'=== uvezi-kr-dosije — {mode} — база: {_database_name()} ===')

        known_users = _load_users()
        plan = plan_import(putanja, odeljenje, known_users)
        d = plan['dossiers']
        m = plan['matches']

        # --- извештај ---
        click.echo(f'\nExcel: {plan["xlsx"].name}')
        click.echo(f'Досијеа (стварних записа): {len(d)}')
        sa_userom = sum(1 for x in d for iz in x['izvrsioci'] if iz['user_email'])
        bez_usera = sum(1 for x in d for iz in x['izvrsioci'] if not iz['user_email'])
        sa_period = sum(1 for x in d if x['period_od'])
        mineral = sum(1 for x in d if x['database_name'] == 'mineral')
        click.echo(f'  везаних за збирку (mineral): {mineral} | слободан унос: {len(d) - mineral}')
        click.echo(f'  извршилаца: {sa_userom} поклопљених налога + {bez_usera} слободних имена')
        click.echo(f'  парсиран период: {sa_period}/{len(d)}')
        click.echo(f'\nСлике ({len(plan["images"])} укупно):')
        click.echo(f'  ✓ тачно поклопљене (аутоматски):   {len(m["exact"])}')
        click.echo(f'  ? двосмислене (број → више досијеа): {len(m["ambiguous"])}')
        click.echo(f'  ✗ без поклапања броја:               {len(m["unmatched"])}')
        click.echo(f'  – без ознаке фазе (спољни/шум):      {len(m["no_phase"])}')
        if m['ambiguous']:
            click.echo('\n  Двосмислене (за ручно повезивање):')
            for ime, n, evbs in m['ambiguous']:
                click.echo(f'    {ime}  (број {n} → досијеи {evbs})')
        if m['unmatched']:
            click.echo('\n  Без поклапања (за ручно повезивање):')
            for ime, n in m['unmatched']:
                click.echo(f'    {ime}  (број {n})')

        if not execute:
            click.echo(f'\nDry-run: ништа није уписано. За упис: --execute')
            return

        # --- потврда ---
        db_name = _database_name()
        odgovor = click.prompt(
            f'\nПОТВРДА: за упис откуцај тачно име базе ({db_name!r})',
            default='', show_default=False,
        )
        if odgovor != db_name:
            raise click.ClickException('Име базе се не поклапа — прекид, без измена.')

        autor_email = KONZERVATOR_EMAIL[odeljenje]
        # Хронолошки редослед → смислени евиденциони бројеви по години.
        redosled = sorted(range(len(d)),
                          key=lambda i: (d[i]['period_od'] or date.max))
        idx_to_id, novih, preskocenih = {}, 0, 0

        with get_postgres_connection() as conn, conn.cursor() as cur:
            for i in redosled:
                postoji = _already_imported(cur, d[i])
                if postoji:
                    idx_to_id[i] = postoji
                    preskocenih += 1
                    continue
                dosije_id, evb = _insert_dossier(cur, d[i], autor_email)
                idx_to_id[i] = dosije_id
                novih += 1
            conn.commit()

            slika_dodato = 0
            if not bez_slika:
                for img, idx, faza in m['exact']:
                    if _upload_and_link(cur, img, idx_to_id[idx], faza, autor_email):
                        slika_dodato += 1
                conn.commit()

        click.echo(f'\nУписано нових досијеа: {novih} (прескочено већ увезених: {preskocenih})')
        if not bez_slika:
            click.echo(f'Отпремљено/повезано слика: {slika_dodato} (тачна поклапања)')
        click.echo('Готово.')
